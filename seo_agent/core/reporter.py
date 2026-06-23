"""
Report generator. Produces two deliverables per run:

  1. seo_report_<date>.xlsx
  2. seo_report_<date>.html
"""

from __future__ import annotations
import datetime as dt
import html
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config
import pdfkit
NAVY = "2E5B8F"
ACCENT = "4A7BA8"



def write_pdf(html_path: Path) -> Path:
    pdf_path = html_path.with_suffix(".pdf")

    options = {
        "enable-local-file-access": "",
        "page-size": "A4",
        "encoding": "UTF-8",
        "margin-top": "10mm",
        "margin-bottom": "10mm",
        "margin-left": "10mm",
        "margin-right": "10mm",
    }

    pdfkit.from_file(str(html_path), str(pdf_path), options=options)
    return pdf_path




def _today() -> str:
    return dt.date.today().isoformat()


def write_excel(rows: list[dict], out_dir: Path | None = None) -> Path:
    out_dir = out_dir or config.REPORTS_DIR
    path = out_dir / f"seo_report_{_today()}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "SEO Action Plan"

    headers = [
        "Group", "Keyword", "Volume", "Difficulty", "Intent",
        "Our Current Rank", "Top Competitor", "Target Page",
        "Word Gap", "Kw in Title?", "Kw in H1?",
        "Suggested Title Tag", "Suggested Meta Description",
        "Priority Actions",
    ]

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color="FFFFFF", bold=True, name="Calibri")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_num, row in enumerate(rows, 2):
        rec = row.get("recommendation", {}) or {}

        values = [
            row.get("group", ""),
            row.get("keyword", ""),
            row.get("volume", ""),
            row.get("difficulty", ""),
            row.get("intent", ""),
            row.get("our_rank", "not ranked"),
            row.get("top_company", ""),
            row.get("target_page", ""),
            row.get("word_gap", ""),
            "Yes" if row.get("kw_in_title") else "No",
            "Yes" if row.get("kw_in_h1") else "No",
            rec.get("title_tag", ""),
            rec.get("meta_description", ""),
            " | ".join(rec.get("priority_actions", [])),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = Font(name="Calibri")
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [14, 34, 9, 10, 14, 14, 16, 28, 9, 11, 10, 40, 48, 50]

    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)

    return path


def write_html(
    rows: list[dict],
    trend_data: list[dict] | None = None,
    out_dir: Path | None = None,
) -> Path:
    out_dir = out_dir or config.REPORTS_DIR
    path = out_dir / f"seo_report_{_today()}.html"

    def esc(value) -> str:
        return html.escape(str(value)) if value is not None else ""

    total_keywords = len(rows)

    not_ranked = sum(
        1 for row in rows
        if not row.get("our_rank")
    )

    avg_word_gap = (
        int(sum(abs(row.get("word_gap", 0) or 0) for row in rows) / len(rows))
        if rows else 0
    )

    top_keyword = max(
        rows,
        key=lambda row: row.get("volume", 0) or 0,
        default={},
    ).get("keyword", "N/A")

    trend_rows = ""

    if trend_data:
      trend_rows = "".join(
        f"""
        <tr>
          <td>{esc(t.get('keyword'))}</td>
          <td>{esc(t.get('previous_rank'))}</td>
          <td>{esc(t.get('current_rank'))}</td>
          <td>{esc(t.get('trend'))}</td>
        </tr>
        """
        for t in trend_data
    )

    gap_rows = "".join(
        f"""
        <tr>
          <td>{esc(row.get('keyword'))}</td>
          <td>{esc(row.get('our_rank') or 'Not ranked')}</td>
          <td>{esc(abs(row.get('word_gap', 0) or 0))}</td>
          <td>{'Yes' if row.get('kw_in_title') else 'No'}</td>
          <td>{'Yes' if row.get('kw_in_h1') else 'No'}</td>
          <td>{esc(row.get('top_company'))}</td>
          <td>{esc(row.get('rankability_score'))}</td> 
          <td>{esc(row.get('priority'))}</td>
        </tr>
        """
        for row in rows
    )

    benchmark_rows = "".join(
    f"""
    <tr>
      <td>{esc(row.get('keyword'))}</td>
      <td>{esc(row.get('jypra_word_count'))}</td>
      <td>{esc(row.get('competitor_avg_word_count'))}</td>
      <td>{esc(abs(row.get('word_gap', 0) or 0))}</td>
      <td>{esc(row.get('jypra_h2_count'))}</td>
      <td>{esc(row.get('competitor_avg_h2_count'))}</td>
    </tr>
    """
    for row in rows
)

    blocks = []

    for row in rows:
        rec = row.get("recommendation", {}) or {}

        faqs = "".join(
            f"<li><b>{esc(faq.get('q'))}</b><br>{esc(faq.get('a'))}</li>"
            for faq in rec.get("faqs", [])
        )

        actions = "".join(
            f"<li>{esc(action)}</li>"
            for action in rec.get("priority_actions", [])
        )

        h2s = "".join(
            f"<li>{esc(h2)}</li>"
            for h2 in rec.get("h2_outline", [])
        )

        competitor_topics = ", ".join(
            esc(term)
            for term in row.get("missing_terms", [])
        )

        seo_topics = "".join(
          f"""
          <div class="topic-card">
            <h4>{esc(topic.get('topic_name'))}</h4>
            <p><b>Why it matters:</b> {esc(topic.get('why_it_matters'))}</p>
            <p><b>Recommended content:</b> {esc(topic.get('recommended_content'))}</p>
            <p><b>Suggested H2:</b> {esc(topic.get('suggested_h2'))}</p>
            <p><b>Recommended word count:</b> {esc(topic.get('recommended_word_count'))}</p>
            <p><b>Ranking benefit:</b> {esc(topic.get('ranking_benefit'))}</p>
          </div>
          """
          for topic in rec.get("seo_topics", [])
)
        topic_content = "".join(
         f"""
         <div class="topic-card">
           <h4>{esc(item.get('topic'))}</h4>
           <p><b>Why it matters:</b> {esc(item.get('why_it_matters'))}</p>
           <p><b>Recommended content:</b></p>
           <p>{esc(item.get('content'))}</p>
           <p><b>Recommended word count:</b> {esc(item.get('recommended_word_count'))}</p>
           <p><b>Ranking benefit:</b> {esc(item.get('ranking_benefit'))}</p>
          </div>
          """
          for item in rec.get("topic_content", [])
)

        blocks.append(f"""
        <div class="card">
          <h2>{esc(row.get('keyword'))}
            <span class="badge">{esc(row.get('group'))}</span>
            <span class="badge">vol {esc(row.get('volume'))}</span>
            <span class="badge">diff {esc(row.get('difficulty'))}</span>
          </h2>

          <p class="meta">
            Current rank: <b>{esc(row.get('our_rank') or 'not ranked')}</b>
            &nbsp;|&nbsp; Top competitor: <b>{esc(row.get('top_company'))}</b>
            &nbsp;|&nbsp; Target page: <code>{esc(row.get('target_page'))}</code>
          </p>

          <h4>Suggested title tag</h4>
          <p>{esc(rec.get('title_tag'))}</p>

          <h4>Suggested meta description</h4>
          <p>{esc(rec.get('meta_description'))}</p>

          <h4>Suggested H1</h4>
          <p>{esc(rec.get('h1'))}</p>

          <h4>Section outline (H2)</h4>
          <ul>{h2s}</ul>

          <h4>Content brief</h4>
          <p>{esc(rec.get('content_brief'))}</p>

         
          <h4>Competitor topics to consider</h4>
          <p class="terms">{competitor_topics}</p>
          <h4>AI Recommended SEO Topic Content</h4>
          {seo_topics}

          <h4>FAQ block (schema-ready)</h4>
          <ul>{faqs}</ul>

          <h4>Priority actions</h4>
          <ol>{actions}</ol>
        </div>
        """)

    doc = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>SEO Action Plan — {esc(config.OWN_BRAND)} — {_today()}</title>

<style>
 body {{
   font-family: Calibri, Arial, sans-serif;
   color: #1f2a37;
   max-width: 1100px;
   margin: 0 auto;
   padding: 32px;
   line-height: 1.5;
   background: #f7f9fc;
 }}

 h1 {{
   color: #{NAVY};
 }}

 h2 {{
   color: #{NAVY};
   margin: 0 0 6px;
 }}

 h3 {{
   color: #{NAVY};
   margin-top: 18px;
 }}

 h4 {{
   color: #{ACCENT};
   margin: 14px 0 2px;
   font-size: .92em;
   text-transform: uppercase;
   letter-spacing: .4px;
 }}

 .card {{
   background: #fff;
   border: 1px solid #e3e9f2;
   border-left: 5px solid #{NAVY};
   border-radius: 10px;
   padding: 20px 24px;
   margin: 18px 0;
   box-shadow: 0 1px 3px rgba(0,0,0,.05);
 }}

 .badge {{
   background: #{ACCENT};
   color: #fff;
   font-size: .55em;
   padding: 3px 8px;
   border-radius: 20px;
   vertical-align: middle;
   margin-left: 6px;
   font-weight: 600;
 }}

 .meta {{
   color: #5b6675;
   font-size: .9em;
 }}

 code {{
   background: #eef3fb;
   padding: 1px 5px;
   border-radius: 4px;
 }}

 .terms {{
   color: #7a4d00;
   font-style: italic;
 }}

 .lead {{
   background: #fff;
   border-radius: 10px;
   padding: 18px 24px;
   border: 1px solid #e3e9f2;
   margin-bottom: 20px;
 }}

 .summary-grid {{
   display: grid;
   grid-template-columns: repeat(2, minmax(0, 1fr));
   gap: 12px;
   margin-top: 12px;
 }}

 .summary-box {{
   background: #f7f9fc;
   border: 1px solid #e3e9f2;
   border-radius: 8px;
   padding: 12px;
 }}

 .summary-number {{
   font-size: 24px;
   font-weight: bold;
   color: #{NAVY};
 }}

 table {{
   width: 100%;
   border-collapse: collapse;
   margin-top: 12px;
 }}

 th, td {{
   border: 1px solid #e3e9f2;
   padding: 8px;
   text-align: left;
   vertical-align: top;
 }}

 th {{
   background: #{NAVY};
   color: white;
 }}
</style>
</head>

<body>

<h1>SEO Action Plan — {esc(config.OWN_BRAND)}</h1>

<div class="lead">
  <b>Generated:</b> {_today()} &nbsp;|&nbsp;
  <b>Keywords analysed:</b> {len(rows)} &nbsp;|&nbsp;
  <b>Domain:</b> {esc(config.OWN_DOMAIN)}

  <hr>

  <h2>Executive Summary</h2>

  <div class="summary-grid">
    <div class="summary-box">
      <div class="summary-number">{total_keywords}</div>
      <div>Total Keywords Analysed</div>
    </div>

    <div class="summary-box">
      <div class="summary-number">{not_ranked}</div>
      <div>Keywords Not Ranking</div>
    </div>

    <div class="summary-box">
      <div class="summary-number">{avg_word_gap}</div>
      <div>Average Word Gap</div>
    </div>

    <div class="summary-box">
      <div class="summary-number">{esc(top_keyword)}</div>
      <div>Highest Volume Keyword</div>
    </div>
  </div>

  <h3>Key Recommendation</h3>
  <p>
    Increase content depth, improve H2 coverage, add FAQ sections,
    and include competitor topic coverage naturally to improve ranking potential.
  </p>

  <p>
    Tip: open this file in Chrome and "Print → Save as PDF" to share with your lead.
  </p>
</div>

<div class="card">
  <h2>Ranking Trend Dashboard</h2>
  <table>
    <tr>
      <th>Keyword</th>
      <th>Previous Rank</th>
      <th>Current Rank</th>
      <th>Trend</th>
    </tr>
    {trend_rows}
  </table>
</div>

<div class="card">
  <h2>Competitor Benchmarking Dashboard</h2>

  <table>
    <tr>
      <th>Keyword</th>
      <th>Jypra Word Count</th>
      <th>Competitor Avg Word Count</th>
      <th>Word Gap</th>
      <th>Jypra H2 Count</th>
      <th>Competitor Avg H2 Count</th>
    </tr>

    {benchmark_rows}

  </table>
</div>

<div class="card">
  <h2>Gap Analysis Table</h2>
  <table>
    <tr>
      <th>Keyword</th>
      <th>Current Rank</th>
      <th>Word Gap</th>
      <th>Keyword in Title</th>
      <th>Keyword in H1</th>
      <th>Top Competitor</th>
      <th>Rankability Score</th>
      <th>Priority</th>
    </tr>
    {gap_rows}
  </table>
</div>

{''.join(blocks)}

</body>
</html>
"""

    path.write_text(doc, encoding="utf-8")
    return path